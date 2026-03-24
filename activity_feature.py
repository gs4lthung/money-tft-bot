from __future__ import annotations

import asyncio
import io
import re
import sqlite3
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Literal, Optional

import discord
from discord import app_commands
from discord.ext import commands
from openpyxl import Workbook
from openpyxl.styles import Font


CHANNEL_MENTION_REGEX = re.compile(r"^<#(\d+)>$")


def register_activity_feature(
    bot: commands.Bot,
    logger,
    activity_db_path: Path,
    owner_user_id: int,
    period_values: set[str],
    metric_values: set[str],
) -> None:
    export_scan_lock = asyncio.Lock()
    last_export_scan_at: dict[int, datetime] = {}
    export_scan_cooldown_seconds = 180
    inactive_page_size = 20

    def init_activity_db() -> None:
        with sqlite3.connect(activity_db_path) as conn:
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS user_activity (
                    guild_id INTEGER NOT NULL,
                    user_id INTEGER NOT NULL,
                    username TEXT NOT NULL,
                    chat_count INTEGER NOT NULL DEFAULT 0,
                    attack_count INTEGER NOT NULL DEFAULT 0,
                    last_active TEXT,
                    PRIMARY KEY (guild_id, user_id)
                )
                """
            )
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS user_activity_events (
                    guild_id INTEGER NOT NULL,
                    channel_id INTEGER NOT NULL DEFAULT 0,
                    user_id INTEGER NOT NULL,
                    username TEXT NOT NULL,
                    event_type TEXT NOT NULL,
                    event_count INTEGER NOT NULL,
                    created_at TEXT NOT NULL
                )
                """
            )
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS processed_messages (
                    message_id INTEGER PRIMARY KEY,
                    guild_id INTEGER NOT NULL,
                    channel_id INTEGER NOT NULL,
                    user_id INTEGER NOT NULL,
                    username TEXT NOT NULL,
                    created_at TEXT NOT NULL
                )
                """
            )
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS channel_scan_state (
                    guild_id INTEGER NOT NULL,
                    channel_id INTEGER NOT NULL,
                    last_message_id INTEGER NOT NULL DEFAULT 0,
                    last_scanned TEXT NOT NULL,
                    PRIMARY KEY (guild_id, channel_id)
                )
                """
            )

            event_columns = {row[1] for row in conn.execute("PRAGMA table_info(user_activity_events)").fetchall()}
            if "channel_id" not in event_columns:
                conn.execute("ALTER TABLE user_activity_events ADD COLUMN channel_id INTEGER NOT NULL DEFAULT 0")

            conn.commit()

    def is_owner_user(user_id: int) -> bool:
        return owner_user_id != 0 and user_id == owner_user_id

    def record_activity(
        guild_id: int,
        channel_id: int,
        user_id: int,
        username: str,
        chat_increment: int = 0,
        attack_increment: int = 0,
    ) -> None:
        now_iso = datetime.now(timezone.utc).isoformat()
        with sqlite3.connect(activity_db_path) as conn:
            conn.execute(
                """
                INSERT INTO user_activity (guild_id, user_id, username, chat_count, attack_count, last_active)
                VALUES (?, ?, ?, ?, ?, ?)
                ON CONFLICT(guild_id, user_id) DO UPDATE SET
                    username = excluded.username,
                    chat_count = user_activity.chat_count + excluded.chat_count,
                    attack_count = user_activity.attack_count + excluded.attack_count,
                    last_active = excluded.last_active
                """,
                (guild_id, user_id, username, chat_increment, attack_increment, now_iso),
            )

            if chat_increment > 0:
                conn.execute(
                    """
                    INSERT INTO user_activity_events
                    (guild_id, channel_id, user_id, username, event_type, event_count, created_at)
                    VALUES (?, ?, ?, ?, 'chat', ?, ?)
                    """,
                    (guild_id, channel_id, user_id, username, chat_increment, now_iso),
                )

            if attack_increment > 0:
                conn.execute(
                    """
                    INSERT INTO user_activity_events
                    (guild_id, channel_id, user_id, username, event_type, event_count, created_at)
                    VALUES (?, ?, ?, ?, 'attack', ?, ?)
                    """,
                    (guild_id, channel_id, user_id, username, attack_increment, now_iso),
                )

            conn.commit()

    def record_chat_events(
        events: list[tuple[int, int, int, int, str, str]],
    ) -> int:
        if not events:
            return 0

        # Aggregate unseen messages to keep DB writes small and avoid double counting.
        aggregated: dict[tuple[int, int, int], tuple[str, int, str]] = {}
        inserted_messages = 0

        def _write_events() -> int:
            nonlocal inserted_messages
            with sqlite3.connect(activity_db_path) as conn:
                for guild_id, channel_id, message_id, user_id, username, created_at in events:
                    cursor = conn.execute(
                        """
                        INSERT OR IGNORE INTO processed_messages
                        (message_id, guild_id, channel_id, user_id, username, created_at)
                        VALUES (?, ?, ?, ?, ?, ?)
                        """,
                        (message_id, guild_id, channel_id, user_id, username, created_at),
                    )
                    if cursor.rowcount == 0:
                        continue

                    inserted_messages += 1
                    key = (guild_id, channel_id, user_id)
                    existing = aggregated.get(key)
                    if existing is None:
                        aggregated[key] = (username, 1, created_at)
                        continue

                    existing_username, existing_count, existing_last_active = existing
                    latest_active = created_at if created_at > existing_last_active else existing_last_active
                    latest_username = username if created_at >= existing_last_active else existing_username
                    aggregated[key] = (latest_username, existing_count + 1, latest_active)

                for (guild_id, channel_id, user_id), (username, chat_increment, last_active) in aggregated.items():
                    conn.execute(
                        """
                        INSERT INTO user_activity (guild_id, user_id, username, chat_count, attack_count, last_active)
                        VALUES (?, ?, ?, ?, ?, ?)
                        ON CONFLICT(guild_id, user_id) DO UPDATE SET
                            username = excluded.username,
                            chat_count = user_activity.chat_count + excluded.chat_count,
                            attack_count = user_activity.attack_count + excluded.attack_count,
                            last_active = excluded.last_active
                        """,
                        (guild_id, user_id, username, chat_increment, 0, last_active),
                    )
                    conn.execute(
                        """
                        INSERT INTO user_activity_events
                        (guild_id, channel_id, user_id, username, event_type, event_count, created_at)
                        VALUES (?, ?, ?, ?, 'chat', ?, ?)
                        """,
                        (guild_id, channel_id, user_id, username, chat_increment, last_active),
                    )

                conn.commit()

            return inserted_messages

        try:
            _write_events()
        except sqlite3.OperationalError as exc:
            # Handles startup race where message events arrive before tables are initialized.
            if "no such table" in str(exc).lower():
                init_activity_db()
                inserted_messages = 0
                aggregated.clear()
                _write_events()
            else:
                raise

        return inserted_messages

    def period_start_iso(period: str) -> Optional[str]:
        now = datetime.now(timezone.utc)
        if period == "all":
            return None
        if period == "day":
            return now.replace(hour=0, minute=0, second=0, microsecond=0).isoformat()
        if period == "month":
            return now.replace(day=1, hour=0, minute=0, second=0, microsecond=0).isoformat()
        raise ValueError(f"Unsupported period: {period}")

    def order_expr(metric: str) -> str:
        if metric == "chat":
            return "chat_count DESC, total_count DESC, attack_count DESC"
        if metric == "attack":
            return "attack_count DESC, total_count DESC, chat_count DESC"
        return "total_count DESC, chat_count DESC, attack_count DESC"

    def get_top_activity_rows(
        guild_id: int,
        limit: int,
        period: str = "all",
        metric: str = "total",
        channel_id: Optional[int] = None,
    ) -> list[tuple[str, int, int, int, str]]:
        sort_by = order_expr(metric)

        with sqlite3.connect(activity_db_path) as conn:
            if period == "all" and channel_id is None:
                rows = conn.execute(
                    f"""
                    SELECT
                        username,
                        chat_count,
                        attack_count,
                        (chat_count + attack_count) AS total_count,
                        COALESCE(last_active, '')
                    FROM user_activity
                    WHERE guild_id = ?
                    ORDER BY {sort_by}
                    LIMIT ?
                    """,
                    (guild_id, limit),
                ).fetchall()
            else:
                start_iso = period_start_iso(period)
                where_clauses = ["guild_id = ?"]
                params: list[object] = [guild_id]

                if channel_id is not None:
                    where_clauses.append("channel_id = ?")
                    params.append(channel_id)

                if start_iso is not None:
                    where_clauses.append("created_at >= ?")
                    params.append(start_iso)

                where_sql = " AND ".join(where_clauses)
                rows = conn.execute(
                    f"""
                    SELECT
                        username,
                        COALESCE(SUM(CASE WHEN event_type = 'chat' THEN event_count ELSE 0 END), 0) AS chat_count,
                        COALESCE(SUM(CASE WHEN event_type = 'attack' THEN event_count ELSE 0 END), 0) AS attack_count,
                        COALESCE(SUM(event_count), 0) AS total_count,
                        COALESCE(MAX(created_at), '') AS last_active
                    FROM user_activity_events
                    WHERE {where_sql}
                    GROUP BY guild_id, user_id
                    ORDER BY {sort_by}
                    LIMIT ?
                    """,
                    tuple([*params, limit]),
                ).fetchall()

        return [(str(r[0]), int(r[1]), int(r[2]), int(r[3]), str(r[4])) for r in rows]

    def get_activity_rows_for_export(
        guild_id: int,
        period: str = "all",
        channel_id: Optional[int] = None,
    ) -> list[tuple[int, str, int, int, int, str]]:
        with sqlite3.connect(activity_db_path) as conn:
            if period == "all" and channel_id is None:
                rows = conn.execute(
                    """
                    SELECT
                        user_id,
                        username,
                        chat_count,
                        attack_count,
                        (chat_count + attack_count) AS total_count,
                        COALESCE(last_active, '')
                    FROM user_activity
                    WHERE guild_id = ?
                    """,
                    (guild_id,),
                ).fetchall()
            else:
                start_iso = period_start_iso(period)
                where_clauses = ["guild_id = ?"]
                params: list[object] = [guild_id]

                if channel_id is not None:
                    where_clauses.append("channel_id = ?")
                    params.append(channel_id)

                if start_iso is not None:
                    where_clauses.append("created_at >= ?")
                    params.append(start_iso)

                where_sql = " AND ".join(where_clauses)
                rows = conn.execute(
                    f"""
                    SELECT
                        user_id,
                        username,
                        COALESCE(SUM(CASE WHEN event_type = 'chat' THEN event_count ELSE 0 END), 0) AS chat_count,
                        COALESCE(SUM(CASE WHEN event_type = 'attack' THEN event_count ELSE 0 END), 0) AS attack_count,
                        COALESCE(SUM(event_count), 0) AS total_count,
                        COALESCE(MAX(created_at), '') AS last_active
                    FROM user_activity_events
                    WHERE {where_sql}
                    GROUP BY guild_id, user_id
                    """,
                    tuple(params),
                ).fetchall()

        return [(int(r[0]), str(r[1]), int(r[2]), int(r[3]), int(r[4]), str(r[5])) for r in rows]

    def build_activity_excel(guild: discord.Guild, period: str = "all", channel_id: Optional[int] = None) -> io.BytesIO:
        rows = get_activity_rows_for_export(guild.id, period=period, channel_id=channel_id)
        by_user_id: dict[int, tuple[str, int, int, int, str]] = {}

        for user_id, username, chat_count, attack_count, total_count, last_active in rows:
            by_user_id[user_id] = (username, chat_count, attack_count, total_count, last_active)

        for member in guild.members:
            if member.bot:
                continue
            if member.id not in by_user_id:
                by_user_id[member.id] = (str(member), 0, 0, 0, "")
            else:
                _, chat_count, attack_count, total_count, last_active = by_user_id[member.id]
                by_user_id[member.id] = (str(member), chat_count, attack_count, total_count, last_active)

        merged_rows = sorted(
            by_user_id.values(),
            key=lambda r: (-r[3], -r[1], -r[2], r[0].lower()),
        )

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Activity"
        sheet.append(["Username", "Chat Count", "Attack Count", "Total", "Last Active (UTC)", "Status"])

        red_bold = Font(color="FFFF0000", bold=True)

        for username, chat_count, attack_count, total_count, last_active in merged_rows:
            status = "NO CHAT" if chat_count == 0 else ""
            sheet.append([username, chat_count, attack_count, total_count, last_active, status])
            if chat_count == 0:
                row_index = sheet.max_row
                for col in range(1, 7):
                    sheet.cell(row=row_index, column=col).font = red_bold

        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        return output

    def get_member_activity_list(
        guild: discord.Guild,
        period: str = "all",
        channel_id: Optional[int] = None,
    ) -> list[tuple[int, str, int, int, int, str, str]]:
        rows = get_activity_rows_for_export(guild.id, period=period, channel_id=channel_id)
        by_user_id: dict[int, tuple[str, int, int, int, str]] = {
            user_id: (username, chat_count, attack_count, total_count, last_active)
            for user_id, username, chat_count, attack_count, total_count, last_active in rows
        }

        members: list[tuple[int, str, int, int, int, str, str]] = []
        for member in guild.members:
            if member.bot:
                continue

            role_name = member.top_role.name if member.top_role is not None else "unknown"

            activity = by_user_id.get(member.id)
            if activity is None:
                members.append((member.id, str(member), 0, 0, 0, "never", role_name))
                continue

            _, chat_count, attack_count, total_count, last_active = activity
            members.append((member.id, str(member), chat_count, attack_count, total_count, last_active or "never", role_name))

        members.sort(key=lambda x: x[1].lower())
        return members

    def format_vietnamese_datetime(raw_value: str) -> str:
        value = (raw_value or "").strip()
        if not value or value.lower() == "never":
            return "chua co"

        try:
            dt = datetime.fromisoformat(value)
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=timezone.utc)
            vn_tz = timezone(timedelta(hours=7))
            return dt.astimezone(vn_tz).strftime("%d/%m/%Y %H:%M")
        except ValueError:
            return value

    def build_member_activity_embed(
        members: list[tuple[int, str, int, int, int, str, str]],
        page_index: int,
        period: str,
        channel_label: str,
        selected_user_id: Optional[int] = None,
        show_inactive_only: bool = False,
        total_members: Optional[int] = None,
        sort_label: str = "Name A-Z",
    ) -> discord.Embed:
        total = len(members)
        if total == 0:
            filter_text = "inactive only" if show_inactive_only else "all members"
            return discord.Embed(
                title=f"Member Activity ({period.upper()} | {channel_label})",
                description=f"No users found for {filter_text}.",
                color=discord.Color.orange(),
                timestamp=datetime.now(timezone.utc),
            )

        total_pages = max(1, (total + inactive_page_size - 1) // inactive_page_size)
        safe_page = max(0, min(page_index, total_pages - 1))
        start = safe_page * inactive_page_size
        end = min(start + inactive_page_size, total)

        lines = []
        for idx, (user_id, username, chat_count, _attack_count, _total_count_row, last_active, role_name) in enumerate(
            members[start:end],
            start=start + 1,
        ):
            marker = " [selected]" if selected_user_id == user_id else ""
            last_active_vn = format_vietnamese_datetime(last_active)
            lines.append(
                f"{idx}. <@{user_id}> ({username}) | role={role_name} | chat={chat_count} "
                f"| last_active={last_active_vn}{marker}"
            )

        filter_text = "INACTIVE ONLY" if show_inactive_only else "ALL MEMBERS"
        embed = discord.Embed(
            title=f"Member Activity ({period.upper()} | {channel_label})",
            description="\n".join(lines),
            color=discord.Color.blurple(),
            timestamp=datetime.now(timezone.utc),
        )
        inactive_count = sum(1 for _, _, chat_count, _, _, _, _ in members if chat_count == 0)
        footer_total = total_members if total_members is not None else total
        embed.set_footer(
            text=(
                f"Page {safe_page + 1}/{total_pages} | Filter: {filter_text} | "
                f"Sort: {sort_label} | Showing {total}/{footer_total} | Inactive in view: {inactive_count}"
            )
        )
        return embed

    class ActivityMembersView(discord.ui.View):
        def __init__(
            self,
            author_id: int,
            guild: discord.Guild,
            members: list[tuple[int, str, int, int, int, str, str]],
            period: str,
            channel_label: str,
        ) -> None:
            super().__init__(timeout=300)
            self.author_id = author_id
            self.guild = guild
            self.members = members
            self.period = period
            self.channel_label = channel_label
            self.page_index = 0
            self.selected_user_id: Optional[int] = None
            self.show_inactive_only = False
            self._sort_modes: list[tuple[str, str]] = [
                ("name_asc", "Name A-Z"),
                ("chat_desc", "Chat High-Low"),
            ]
            self.sort_mode_index = 0

            self.select_user_menu = discord.ui.Select(
                placeholder="Select a user",
                min_values=1,
                max_values=1,
                options=[discord.SelectOption(label="No users", value="0")],
                disabled=True,
            )
            self.select_user_menu.callback = self.select_user_callback
            self.add_item(self.select_user_menu)
            self._refresh_components()

        def _current_sort_mode(self) -> str:
            return self._sort_modes[self.sort_mode_index][0]

        def _current_sort_label(self) -> str:
            return self._sort_modes[self.sort_mode_index][1]

        def _visible_members(self) -> list[tuple[int, str, int, int, int, str, str]]:
            if self.show_inactive_only:
                filtered = [row for row in self.members if row[2] == 0]
            else:
                filtered = list(self.members)

            sort_mode = self._current_sort_mode()
            if sort_mode == "chat_desc":
                filtered.sort(key=lambda row: (-row[2], row[1].lower()))
            else:
                filtered.sort(key=lambda row: row[1].lower())

            return filtered

        def _refresh_components(self) -> None:
            filtered_members = self._visible_members()
            total_pages = max(1, (len(filtered_members) + inactive_page_size - 1) // inactive_page_size)
            safe_page = max(0, min(self.page_index, total_pages - 1))
            self.page_index = safe_page
            start = safe_page * inactive_page_size
            end = min(start + inactive_page_size, len(filtered_members))
            current_page_user_ids = {user_id for user_id, _, _, _, _, _, _ in filtered_members[start:end]}

            options: list[discord.SelectOption] = []
            for idx, (user_id, username, chat_count, _, _, last_active, role_name) in enumerate(
                filtered_members[start:end],
                start=start + 1,
            ):
                numbered_label = f"{idx}. {username}"
                last_active_vn = format_vietnamese_datetime(last_active)
                options.append(
                    discord.SelectOption(
                        label=numbered_label[:100],
                        description=f"role={role_name} | chat={chat_count} | last_active={last_active_vn}"[:100],
                        value=str(user_id),
                        default=self.selected_user_id == user_id,
                    )
                )

            if self.selected_user_id is not None and self.selected_user_id not in current_page_user_ids:
                self.selected_user_id = None

            if options:
                self.select_user_menu.options = options
                self.select_user_menu.disabled = False
            else:
                self.select_user_menu.options = [discord.SelectOption(label="No users", value="0")]
                self.select_user_menu.disabled = True

            self.prev_button.disabled = safe_page <= 0
            self.next_button.disabled = safe_page >= total_pages - 1
            self.kick_button.disabled = self.selected_user_id is None
            self.toggle_filter_button.label = "Show All Members" if self.show_inactive_only else "Show Inactive Only"
            self.toggle_filter_button.style = (
                discord.ButtonStyle.success if self.show_inactive_only else discord.ButtonStyle.primary
            )
            self.sort_button.label = f"Sort: {self._current_sort_label()}"

        async def interaction_check(self, interaction: discord.Interaction) -> bool:
            if interaction.user.id != self.author_id:
                await interaction.response.send_message("[x] Only the command caller can use these controls.", ephemeral=True)
                return False
            return True

        async def select_user_callback(self, interaction: discord.Interaction) -> None:
            selected_value = self.select_user_menu.values[0]
            self.selected_user_id = None if selected_value == "0" else int(selected_value)
            self._refresh_components()
            filtered_members = self._visible_members()
            await interaction.response.edit_message(
                embed=build_member_activity_embed(
                    filtered_members,
                    self.page_index,
                    self.period,
                    self.channel_label,
                    selected_user_id=self.selected_user_id,
                    show_inactive_only=self.show_inactive_only,
                    total_members=len(self.members),
                    sort_label=self._current_sort_label(),
                ),
                view=self,
            )

        @discord.ui.button(label="Prev", style=discord.ButtonStyle.secondary)
        async def prev_button(self, interaction: discord.Interaction, _: discord.ui.Button) -> None:
            self.page_index = max(0, self.page_index - 1)
            self._refresh_components()
            filtered_members = self._visible_members()
            await interaction.response.edit_message(
                embed=build_member_activity_embed(
                    filtered_members,
                    self.page_index,
                    self.period,
                    self.channel_label,
                    selected_user_id=self.selected_user_id,
                    show_inactive_only=self.show_inactive_only,
                    total_members=len(self.members),
                    sort_label=self._current_sort_label(),
                ),
                view=self,
            )

        @discord.ui.button(label="Next", style=discord.ButtonStyle.secondary)
        async def next_button(self, interaction: discord.Interaction, _: discord.ui.Button) -> None:
            filtered_members = self._visible_members()
            total_pages = max(1, (len(filtered_members) + inactive_page_size - 1) // inactive_page_size)
            self.page_index = min(total_pages - 1, self.page_index + 1)
            self._refresh_components()
            await interaction.response.edit_message(
                embed=build_member_activity_embed(
                    filtered_members,
                    self.page_index,
                    self.period,
                    self.channel_label,
                    selected_user_id=self.selected_user_id,
                    show_inactive_only=self.show_inactive_only,
                    total_members=len(self.members),
                    sort_label=self._current_sort_label(),
                ),
                view=self,
            )

        @discord.ui.button(label="Sort: Name A-Z", style=discord.ButtonStyle.secondary)
        async def sort_button(self, interaction: discord.Interaction, _: discord.ui.Button) -> None:
            self.sort_mode_index = (self.sort_mode_index + 1) % len(self._sort_modes)
            self.page_index = 0
            self.selected_user_id = None
            self._refresh_components()
            filtered_members = self._visible_members()
            await interaction.response.edit_message(
                embed=build_member_activity_embed(
                    filtered_members,
                    self.page_index,
                    self.period,
                    self.channel_label,
                    selected_user_id=self.selected_user_id,
                    show_inactive_only=self.show_inactive_only,
                    total_members=len(self.members),
                    sort_label=self._current_sort_label(),
                ),
                view=self,
            )

        @discord.ui.button(label="Show Inactive Only", style=discord.ButtonStyle.primary)
        async def toggle_filter_button(self, interaction: discord.Interaction, _: discord.ui.Button) -> None:
            self.show_inactive_only = not self.show_inactive_only
            self.page_index = 0
            self.selected_user_id = None
            self._refresh_components()
            filtered_members = self._visible_members()
            await interaction.response.edit_message(
                embed=build_member_activity_embed(
                    filtered_members,
                    self.page_index,
                    self.period,
                    self.channel_label,
                    selected_user_id=self.selected_user_id,
                    show_inactive_only=self.show_inactive_only,
                    total_members=len(self.members),
                    sort_label=self._current_sort_label(),
                ),
                view=self,
            )

        @discord.ui.button(label="Kick Selected", style=discord.ButtonStyle.danger)
        async def kick_button(self, interaction: discord.Interaction, _: discord.ui.Button) -> None:
            if self.selected_user_id is None:
                await interaction.response.send_message("[x] Select a user first.", ephemeral=True)
                return

            selected_row = next((row for row in self.members if row[0] == self.selected_user_id), None)
            if selected_row is None:
                await interaction.response.send_message("[x] Selected user no longer exists in the list.", ephemeral=True)
                return

            if selected_row[2] > 0:
                await interaction.response.send_message(
                    "[x] Kick action in this panel is only for inactive users (chat=0).",
                    ephemeral=True,
                )
                return

            if not interaction.user.guild_permissions.kick_members:
                await interaction.response.send_message("[x] You need Kick Members permission.", ephemeral=True)
                return

            me = self.guild.me
            if me is None or not me.guild_permissions.kick_members:
                await interaction.response.send_message("[x] Bot does not have Kick Members permission.", ephemeral=True)
                return

            member = self.guild.get_member(self.selected_user_id)
            if member is None:
                await interaction.response.send_message("[x] User is not in server anymore.", ephemeral=True)
                return

            if member.top_role >= me.top_role:
                await interaction.response.send_message("[x] Cannot kick: user role is higher/equal than bot role.", ephemeral=True)
                return

            if member.top_role >= interaction.user.top_role and interaction.user.id != self.guild.owner_id:
                await interaction.response.send_message("[x] Cannot kick: user role is higher/equal than your role.", ephemeral=True)
                return

            try:
                await member.kick(reason=f"Inactive cleanup by {interaction.user} ({interaction.user.id})")
            except discord.Forbidden:
                await interaction.response.send_message("[x] Kick failed due to Discord permissions.", ephemeral=True)
                return
            except Exception:
                logger.exception("Failed kicking inactive user=%s guild=%s", self.selected_user_id, self.guild.id)
                await interaction.response.send_message("[x] Kick failed due to an unexpected error.", ephemeral=True)
                return

            self.members = [u for u in self.members if u[0] != self.selected_user_id]
            self.selected_user_id = None
            self._refresh_components()
            filtered_members = self._visible_members()
            await interaction.response.edit_message(
                embed=build_member_activity_embed(
                    filtered_members,
                    self.page_index,
                    self.period,
                    self.channel_label,
                    show_inactive_only=self.show_inactive_only,
                    total_members=len(self.members),
                    sort_label=self._current_sort_label(),
                ),
                view=self,
            )
            await interaction.followup.send("User kicked successfully.", ephemeral=True)

        async def on_timeout(self) -> None:
            self.select_user_menu.disabled = True
            self.prev_button.disabled = True
            self.next_button.disabled = True
            self.sort_button.disabled = True
            self.toggle_filter_button.disabled = True
            self.kick_button.disabled = True

    async def scan_full_guild_history(guild: discord.Guild) -> tuple[int, int, int]:
        me = guild.me
        if me is None:
            return 0, 0, len(guild.text_channels)

        with sqlite3.connect(activity_db_path) as conn:
            known_scan_state = {
                int(row[0]): int(row[1])
                for row in conn.execute(
                    "SELECT channel_id, last_message_id FROM channel_scan_state WHERE guild_id = ?",
                    (guild.id,),
                ).fetchall()
            }

        scanned_total = 0
        added_total = 0
        skipped_channels = 0
        batch: list[tuple[int, int, int, int, str, str]] = []
        batch_size = 500
        since_pause_counter = 0

        for channel in guild.text_channels:
            perms = channel.permissions_for(me)
            if not perms.read_messages or not perms.read_message_history:
                skipped_channels += 1
                continue

            try:
                previous_last_message_id = known_scan_state.get(channel.id)
                latest_seen_message_id = previous_last_message_id or 0
                channel_last_message_id = int(channel.last_message_id or 0)

                # Fast path: if we already scanned this channel up to the current last message,
                # skip history calls entirely.
                if previous_last_message_id and channel_last_message_id and channel_last_message_id <= previous_last_message_id:
                    continue

                # Fast path: empty channel, no need to call history API.
                if channel_last_message_id == 0:
                    continue

                history_kwargs: dict[str, object] = {"limit": None}
                if previous_last_message_id:
                    history_kwargs["after"] = discord.Object(id=previous_last_message_id)

                async for msg in channel.history(**history_kwargs):
                    if msg.author.bot:
                        continue

                    scanned_total += 1
                    since_pause_counter += 1
                    if msg.id > latest_seen_message_id:
                        latest_seen_message_id = msg.id

                    batch.append(
                        (
                            guild.id,
                            channel.id,
                            msg.id,
                            msg.author.id,
                            str(msg.author),
                            msg.created_at.astimezone(timezone.utc).isoformat(),
                        )
                    )

                    if len(batch) >= batch_size:
                        added_total += record_chat_events(batch)
                        batch.clear()

                    # Prevent bursty paging over very large channels.
                    if since_pause_counter >= 250:
                        await asyncio.sleep(1.0)
                        since_pause_counter = 0

                if latest_seen_message_id > 0:
                    now_iso = datetime.now(timezone.utc).isoformat()
                    with sqlite3.connect(activity_db_path) as conn:
                        conn.execute(
                            """
                            INSERT INTO channel_scan_state (guild_id, channel_id, last_message_id, last_scanned)
                            VALUES (?, ?, ?, ?)
                            ON CONFLICT(guild_id, channel_id) DO UPDATE SET
                                last_message_id = excluded.last_message_id,
                                last_scanned = excluded.last_scanned
                            """,
                            (guild.id, channel.id, latest_seen_message_id, now_iso),
                        )
                        conn.commit()

                # Smooth out request bursts across channels to reduce 429 responses.
                await asyncio.sleep(1.0)
            except Exception:
                skipped_channels += 1
                logger.exception("Failed export scan for channel=%s guild=%s", channel.id, guild.id)

        if batch:
            added_total += record_chat_events(batch)

        return scanned_total, added_total, skipped_channels

    def has_any_all_time_activity(guild_id: int) -> bool:
        with sqlite3.connect(activity_db_path) as conn:
            row = conn.execute(
                "SELECT COUNT(*) FROM user_activity WHERE guild_id = ?",
                (guild_id,),
            ).fetchone()
        return bool(row and int(row[0]) > 0)

    async def bootstrap_all_time_activity_if_needed(guild: discord.Guild) -> tuple[bool, int, int, int]:
        if has_any_all_time_activity(guild.id):
            return False, 0, 0, 0

        if export_scan_lock.locked():
            return False, 0, 0, 0

        async with export_scan_lock:
            scanned_total, added_total, skipped_channels = await scan_full_guild_history(guild)
            last_export_scan_at[guild.id] = datetime.now(timezone.utc)
            return True, scanned_total, added_total, skipped_channels

    @bot.listen("on_ready")
    async def activity_on_ready() -> None:
        init_activity_db()

    # Initialize immediately as well to avoid on_message race during startup.
    init_activity_db()

    @bot.listen("on_message")
    async def activity_on_message(message: discord.Message) -> None:
        if message.author.bot:
            return

        if message.guild is not None:
            try:
                record_chat_events(
                    [
                        (
                            message.guild.id,
                            message.channel.id,
                            message.id,
                            message.author.id,
                            str(message.author),
                            message.created_at.astimezone(timezone.utc).isoformat(),
                        )
                    ]
                )
            except Exception:
                logger.exception("Failed to record chat activity for user=%s", message.author.id)

        await bot.process_commands(message)

    @bot.listen("on_attack_event")
    async def activity_on_attack_event(guild_id: int, channel_id: int, user_id: int, username: str) -> None:
        try:
            record_activity(
                guild_id=guild_id,
                channel_id=channel_id,
                user_id=user_id,
                username=username,
                attack_increment=1,
            )
        except Exception:
            logger.exception("Failed to record attack activity for user=%s in guild=%s", user_id, guild_id)

    @bot.command(name="activity_export")
    async def activity_export_prefix(ctx: commands.Context) -> None:
        if ctx.guild is None:
            await ctx.reply("[x] This command can only be used in a server.")
            return

        if not ctx.author.guild_permissions.kick_members:
            await ctx.reply("[x] You need Kick Members permission to run this command.")
            return

        if export_scan_lock.locked():
            await ctx.reply("[x] Export scan is already running. Please wait for it to finish.")
            return

        now = datetime.now(timezone.utc)
        last_run = last_export_scan_at.get(ctx.guild.id)
        if last_run is not None:
            elapsed = (now - last_run).total_seconds()
            if elapsed < export_scan_cooldown_seconds:
                wait_seconds = int(export_scan_cooldown_seconds - elapsed)
                await ctx.reply(f"[x] Please wait {wait_seconds}s before running export again.")
                return

        await ctx.reply("Scanning channels before showing member activity list...")

        try:
            async with export_scan_lock:
                scanned_total, added_total, skipped_channels = await scan_full_guild_history(ctx.guild)
                last_export_scan_at[ctx.guild.id] = datetime.now(timezone.utc)

            members = get_member_activity_list(ctx.guild, period="all", channel_id=None)
            if not members:
                await ctx.send("No members found after scan.")
                return

            await ctx.send(
                "Scan complete. "
                f"scanned={scanned_total}, added={added_total}, "
                f"ignored_duplicates={max(0, scanned_total - added_total)}, skipped_channels={skipped_channels}"
            )

            channel_label = "ALL CHANNELS"
            view = ActivityMembersView(
                author_id=ctx.author.id,
                guild=ctx.guild,
                members=members,
                period="all",
                channel_label=channel_label,
            )
            await ctx.send(
                embed=build_member_activity_embed(
                    members,
                    0,
                    "all",
                    channel_label,
                    show_inactive_only=False,
                    total_members=len(members),
                ),
                view=view,
            )
        except Exception:
            logger.exception("Failed to export activity for guild=%s", ctx.guild.id)
            await ctx.reply("[x] Failed to load member activity list.")

    @bot.tree.command(name="activity_export", description="Show member activity list with inactive toggle and kick controls")
    async def activity_export_slash(
        interaction: discord.Interaction,
    ) -> None:
        if interaction.guild is None:
            await interaction.response.send_message("[x] This command can only be used in a server.", ephemeral=True)
            return

        if not interaction.user.guild_permissions.kick_members:
            await interaction.response.send_message(
                "[x] You need Kick Members permission to run this command.",
                ephemeral=True,
            )
            return

        if export_scan_lock.locked():
            await interaction.response.send_message(
                "[x] Export scan is already running. Please wait for it to finish.",
                ephemeral=True,
            )
            return

        now = datetime.now(timezone.utc)
        last_run = last_export_scan_at.get(interaction.guild.id)
        if last_run is not None:
            elapsed = (now - last_run).total_seconds()
            if elapsed < export_scan_cooldown_seconds:
                wait_seconds = int(export_scan_cooldown_seconds - elapsed)
                await interaction.response.send_message(
                    f"[x] Please wait {wait_seconds}s before running export again.",
                    ephemeral=True,
                )
                return

        try:
            await interaction.response.defer(ephemeral=True, thinking=True)

            async with export_scan_lock:
                scanned_total, added_total, skipped_channels = await scan_full_guild_history(interaction.guild)
                last_export_scan_at[interaction.guild.id] = datetime.now(timezone.utc)

            members = get_member_activity_list(interaction.guild, period="all", channel_id=None)
            if not members:
                await interaction.followup.send("No members found after scan.", ephemeral=True)
                return

            channel_label = "ALL CHANNELS"
            view = ActivityMembersView(
                author_id=interaction.user.id,
                guild=interaction.guild,
                members=members,
                period="all",
                channel_label=channel_label,
            )
            await interaction.followup.send(
                "Member activity list ready after scan. "
                f"scanned={scanned_total}, added={added_total}, "
                f"ignored_duplicates={max(0, scanned_total - added_total)}, skipped_channels={skipped_channels}",
                embed=build_member_activity_embed(
                    members,
                    0,
                    "all",
                    channel_label,
                    show_inactive_only=False,
                    total_members=len(members),
                ),
                view=view,
                ephemeral=True,
            )
        except Exception:
            logger.exception("Failed slash activity export for guild=%s", interaction.guild_id)
            if interaction.response.is_done():
                await interaction.followup.send("[x] Failed to load member activity list.", ephemeral=True)
            else:
                await interaction.response.send_message("[x] Failed to load member activity list.", ephemeral=True)
